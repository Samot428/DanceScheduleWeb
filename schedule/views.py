from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.urls import reverse
from .models import UploadedScheduleFile
from main.models import DancersAvailability, Day, Dancer, Group, Couple, Trainer, TrainerDayAvailability, GroupLesson
from TrainerClubs.models import Club
from sheet.models import SheetCell
from .reading_excel_func import read_dancers_availability
from .reading_template_func import read_dancers_availability_from_template
from openpyxl import load_workbook
import os
import logging
import tempfile
import requests
from datetime import time as time_make
from collections import defaultdict
from .making_schedule_func import (
    convert_to_min,
    min_to_time_str,
    make_trainers_windows,
    sort_couples_by_group,
    create_schedule as build_schedule,
    create_schedule_with_escalating_pairs,
)
from .making_schedule_func2 import build_schedule as build_schedule2
from django.contrib.auth.decorators import login_required

# Configure logging
logger = logging.getLogger(__name__)

# Constants for scheduling
MIN_AVAILABLE_SLOTS_THRESHOLD = 8
MAX_PAIR_COUNT_DEFAULT = 20
BASE_TIMEOUT = 5
TIMEOUT_PER_COUPLE = 2.5
MAX_TIMEOUT = 60
STALL_THRESHOLD = 5000
PROGRESS_TIMEOUT = 5
# Create your views here.
@login_required
def calendar_view(request):
    return redirect('main:calendar_view')

@login_required
@ensure_csrf_cookie
def schedule_view(request, club_id):
    club = get_object_or_404(Club, id=club_id, club_owner=request.user)
    days = Day.objects.filter(user=request.user, club=club)
    return render(request, 'schedule_view.html', {'days': days, 'club':club})

def validate_excel_format(file_path, user):
    """Validate if the Excel file has the correct format for reading_excel_func.py"""
    try:
        # Check if file exists and is a valid Excel file
        wb = load_workbook(file_path)

        # Check if at least one group sheet exists
        groups = Group.objects.filter(user=user)
        if not groups.exists():
            return False, "No groups defined in the database. Please define groups first." 

        # Validate that at least one group sheet exists in the workbook
        found_valid_sheet = False
        for group in groups:
            if group.name in wb.sheetnames:
                found_valid_sheet = True
                try:
                    # Try to read availability from this sheet
                    day_times, day_dancers_avail = read_dancers_availability(group.name, file_path)

                    # Check if data was extracted
                    if not day_times or not day_dancers_avail:
                        return False, f'Sheet "{group.name}" exists but has no valid data structure.'

                except Exception as e:
                    return False, f'Sheet "{group.name}" format error: {str(e)}'
        if not found_valid_sheet:
            available_sheets = ", ".join(wb.sheetnames)
            expected_sheets = ", ".join([g.name for g in groups])
            return False, f"No valid group sheets found. Expected: {expected_sheets}, Found: {available_sheets}" 
        return True, 'File format is valid'
    except Exception as e:
        return False, f'Error reading file: {str(e)}'

@login_required
def upload_schedule_files(request, club_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)

    files = request.FILES.getlist('files') or []
    if not files:
        return JsonResponse({'status': 'error', 'message': 'No files provided'}, status=400)
    
    club = get_object_or_404(Club, id=club_id)
    uploaded_files_info = []
    for f in files:
        uploaded_file = UploadedScheduleFile.objects.create(
            filename=f.name,
            file=f,
            user=request.user,
            club=club,
        )

        # # Validate the uploaded file format
        # is_valid, error_message = validate_excel_format(uploaded_file.file.path, request.user)

        try:
            response = requests.get(uploaded_file.file.url)
            response.raise_for_status()

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(response.content)
                tmp_path = tmp.name

            is_valid, error_message = validate_excel_format(tmp_path, request.user)
        except Exception as e:
            uploaded_file.file.delete(save=False)
            uploaded_file.delete()
            return JsonResponse({
                'status': 'error',
                'message': f'Error while loading the File: {str(e)}'
            }, status=400)

        if not is_valid:
            # Delete invalid file and return error
            uploaded_file.file.delete(save=False)
            uploaded_file.delete()
            return JsonResponse({
                'status':'error',
                'message':f'File "{f.name}" has invalid format: {error_message}'
            }, status=400)

        uploaded_files_info.append({
            'id': uploaded_file.id,
            'filename': uploaded_file.filename,
            'size': uploaded_file.file.size,
            'uploaded_at': uploaded_file.uploaded_at.strftime('%Y-%m-%d %H:%M'),
            'url': uploaded_file.file.url,  # useful for preview/download
        })

    os.remove(tmp_path)

    return JsonResponse({'status': 'success', 'files': uploaded_files_info, 'club_id':club_id})

@login_required
def get_uploaded_files(request, club_id):
    club = get_object_or_404(Club, id=club_id)
    files = UploadedScheduleFile.objects.filter(club=club, user=request.user).order_by('-uploaded_at')
    files_data = [{
        'id': f.id,
        'filename': f.filename,
        'size': f.file.size,
        'uploaded_at': f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
        'url': f.file.url,
    } for f in files]
    return JsonResponse({'files': files_data})

from django.views.decorators.csrf import csrf_exempt

@login_required
def delete_uploaded_file(request, file_id, club_id):
    if request.method not in ('DELETE', 'POST'):
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)
    club = get_object_or_404(Club, id=club_id)
    f = get_object_or_404(UploadedScheduleFile, id=file_id, user=request.user, club=club)

    f.file.delete(save=False)
    f.delete()
    return JsonResponse({'status': 'success'})

def load_dancers_availability(file_id, user):
    """Load dancer's availability from the uploaded Excel file"""
    # if request.method != 'POST':
    #     return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)
    uploaded_file = UploadedScheduleFile.objects.filter(id=file_id, user=user).first()
    
    try:
        response = requests.get(uploaded_file.file.url)
        response.raise_for_status()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(response.content)
            tmp_path = tmp.name
    except Exception as e:
        return JsonResponse({
            'status':'error',
            'message':f'Erro while loading the File: {str(e)}'
        }, status=400)


    for group in Group.objects.filter(user=user):
        day_times, day_dancers_avail = read_dancers_availability(group.name, tmp_path)
        for day_name, dancers_availability in day_dancers_avail.items():
            day_obj, _ = Day.objects.get_or_create(name=day_name, user=user, defaults={"user": user})
            for dancer_name, availability in dancers_availability.items():
                dancer_obj, _ = Dancer.objects.get_or_create(name=dancer_name)
                DancersAvailability.objects.update_or_create(
                    user=user,
                    dancer=dancer_obj,
                    day=day_obj,
                    defaults={'availability': availability}
                )
    if tmp_path and os.path.exists(tmp_path):
        os.remove(tmp_path)


def parse_couple_names(couple_name):
    """Parse couple name to extract individual dancer names.
    
    Returns:
        tuple: (dancer1_name, dancer2_name) or (dancer_name, None) for single dancers
    """
    name = str(couple_name).strip()
    
    # Try ' & ' separator
    if ' & ' in name:
        parts = name.split(' & ', 1)
        return parts[0].strip(), parts[1].strip()

    # Try 'and' separator first
    if ' and ' in name:
        parts = name.split(' and ', 1)
        return parts[0].strip(), parts[1].strip()
    
    # Try ' a ' separator
    if ' a ' in name:
        parts = name.split(' a ', 1)
        return parts[0].strip(), parts[1].strip()
    
    # Single dancer (not a couple)
    return name, None


def calculate_couple_availability_for_day(couple, day_times_list, dancers_for_day, day_obj=None, with_group_lessons=False):
    """Calculate availability for a couple on a specific day.
    
    Args:
        couple: Couple object
        day_times_list: List of time objects for the day
        dancers_for_day: Dict mapping dancer_name -> list of (time_str, is_available)
        day_obj: Day object (required if with_group_lessons=True)
        with_group_lessons: If True, exclude group lesson times
        
    Returns:
        list: List of (time_str, is_available) tuples
    """
    dancer1_name, dancer2_name = parse_couple_names(couple.name)
    avail = []
    
    # Precompute group lessons if needed
    group_lesson_intervals = []
    if with_group_lessons and day_obj:
        try:
            for gl in day_obj.group_lessons.all():
                gl_groups = gl.groups.all()
                stg = convert_to_min(gl.time_interval_start)
                etg = convert_to_min(gl.time_interval_end)
                
                # Check if couple is in any of these groups
                is_couple_in_group = any(
                    couple in group.couples.all() 
                    for group in gl_groups
                )
                
                if is_couple_in_group:
                    group_lesson_intervals.append((stg, etg))
        except Exception as e:
            # If there's an error, continue without group lesson filtering
            pass
    
    # Calculate availability for each time slot
    for i, time_obj in enumerate(day_times_list):
        time_min = convert_to_min(time_obj)
        time_str = min_to_time_str(time_min)
        
        # Check if both dancers are available
        d1_slots = dancers_for_day.get(dancer1_name, [])
        d2_slots = dancers_for_day.get(dancer2_name, []) if dancer2_name else d1_slots
        
        has_d1 = i < len(d1_slots)
        has_d2 = i < len(d2_slots)
        
        available = bool(has_d1 and has_d2 and d1_slots[i][1] and d2_slots[i][1])
        
        # Exclude group lesson times if requested
        if with_group_lessons and available:
            for stg, etg in group_lesson_intervals:
                if stg <= time_min < etg:
                    available = False
                    break
        
        avail.append((time_str, available))
    
    return avail
   
@login_required
def create_schedule(request, club_id):
    
    """Create schedule for all the days that were inserted in the database by user"""
    if request.method == 'POST':
        try:
            club = get_object_or_404(Club, id=club_id, club_owner=request.user)
            file_id = request.POST.get('file_id')
            sort_couples_by = request.POST.get('sort_by', 'Group Index')
            forday = request.POST.get('days_sort', 'all')
            # uploaded_file = UploadedScheduleFile.objects.get(id=file_id, club=club)
            availability_source = request.POST.get("availability_file")
            use_local_sheet = availability_source == "local"
            tmp_path = None
            if not use_local_sheet:
                uploaded_file = get_object_or_404(UploadedScheduleFile, id=file_id, user=request.user, club=club)
                try:
                    response = requests.get(uploaded_file.file.url)
                    response.raise_for_status()

                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                        tmp.write(response.content)
                        tmp_path = tmp.name

                    logger.info(f'✅ Datei heruntergeladen: {uploaded_file.filename}, Größe: {len(response.content)} bytes')
                    wb_check = load_workbook(tmp_path, read_only=True)
                    logger.info(f'✅ Excel Sheets gefunden: {wb_check.sheetnames}')
                    wb_check.close()
                except Exception as e:
                    return JsonResponse({
                        'status':'error',
                        'message':f'Erro while loading the File: {str(e)}'
                    }, status=400)
            dawt = defaultdict(list)
            cawt = defaultdict(list)
            cawt_with_group_lessons = defaultdict(list)
            
            # Prefetch days to avoid repeated queries
            days_with_group_lessons = Day.objects.filter(user=request.user, club=club).prefetch_related('group_lessons__groups__couples').all()
            days_lookup = {day.name: day for day in days_with_group_lessons}
            
            for group in Group.objects.filter(user=request.user, club=club).prefetch_related('couples').all():
                if use_local_sheet:
                    day_times, dancers_avail = read_dancers_availability_from_template(club_id, group.name)
                else:
                    # Prüfen ob das Sheet existiert, bevor wir es lesen
                    try:
                        wb_check = load_workbook(tmp_path, read_only=True)
                        if group.name not in wb_check.sheetnames:
                            logger.info(f'Sheet "{group.name}" not found in Excel, skipping group.')
                            wb_check.close()
                            continue  # ← diese Gruppe überspringen
                        wb_check.close()
                    except Exception as e:
                        logger.warning(f'Could not check workbook: {e}')
                        continue
                    
                    day_times, dancers_avail = read_dancers_availability(group.name, tmp_path)
                # compute dancers availablity with times in format (time_str, True/False)
                for day in dancers_avail:
                    dancers = {}
                    for d in dancers_avail[day]:
                        avail = []
                        day_times_list = day_times[day]
                        for i in range(len(day_times_list)):
                            time_obj = day_times_list[i]
                            time_min = convert_to_min(time_obj)
                            time_str = min_to_time_str(time_min)
                            avail.append((time_str, dancers_avail[day][d][i]))
                        dancers[d] = avail
                    dawt[day].append(dancers)
                # couples availability with times
                couples = group.couples.all()
                dancers = group.dancers.all() 
                all_schedules = {}
                dawt_for_dancer_usage = dawt.copy()

                for day in day_times:
                    dancers_to_use = []
                    current_day_obj = get_object_or_404(Day, name=day, club=club) 
                    for d in dancers:
                        if d.in_couple:
                            if d.sex == "male":
                                c = get_object_or_404(Couple, man=d)
                            elif d.sex == "female":
                                c = get_object_or_404(Couple, woman=d)
                            if c not in current_day_obj.couples.all() and d in current_day_obj.dancers.all():
                                dancers_to_use.append(d)
                        elif d in current_day_obj.dancers.all():
                            dancers_to_use.append(d)

                    day_couples = {}
                    day_couples1 = {}

                    # Merge all dancers from all groups for this day
                    dancers_for_day = {}
                    if dawt[day]:
                        for group_dancers in dawt[day]:
                            dancers_for_day.update(group_dancers)
                    
                    # Get day object for group lesson checking (use prefetched lookup)
                    day_obj = days_lookup.get(day)
                    day_times_list = day_times[day]
                    
                    for couple in couples:
                        # Calculate availability without group lessons
                        avail = calculate_couple_availability_for_day(
                            couple, day_times_list, dancers_for_day, day_obj, with_group_lessons=False
                        )
                        
                        # Calculate availability with group lessons excluded
                        avail1 = calculate_couple_availability_for_day(
                            couple, day_times_list, dancers_for_day, day_obj, with_group_lessons=True
                        )
                        
                        # Only add couples that have some availability
                        if any(slot[1] for slot in avail):
                            day_couples[couple.name] = avail
                        
                        # Add to group-lesson-aware availability
                        if any(slot[1] for slot in avail1):
                            day_couples1[couple.name] = avail1
                        elif any(slot[1] for slot in avail):
                            # Fallback to regular availability if no slots with group lessons
                            day_couples1[couple.name] = avail
                    for dancer in dancers_to_use:
                        group_lesson_intervals = []
                        try:
                            for gl in current_day_obj.group_lessons.all():
                                gl_groups = gl.groups.all()
                                stg = convert_to_min(gl.time_interval_start)
                                etg = convert_to_min(gl.time_interval_end)

                                dancer_in_group = any(
                                    dancer in group.dancers.all()
                                    for group in gl_groups
                                )

                                if dancer_in_group:
                                    group_lesson_intervals.append((stg, etg))
                        except Exception as e:
                            pass

                        avail = []
                        avail1 = []
                        for i, time_obj in enumerate(day_times_list):
                            time_min = convert_to_min(time_obj)
                            time_str = min_to_time_str(time_min)

                            d_slots = dancers_for_day.get(dancer.name, [])
                            has_d = i < len(d_slots)

                            available = bool(has_d and d_slots[i][1])

                            avail.append((time_str, available))
                            for stg, etg in group_lesson_intervals:
                                if stg <= time_min <= etg:
                                    available = False
                                    break
                            avail1.append((time_str, available))

                        if any(slot[1] for slot in avail):
                            day_couples[dancer.name] = avail
                        if any(slot[1] for slot in avail1):
                            day_couples1[dancer.name] = avail1
                        elif any(slot[1] for slot in avail):
                            day_couples1[dancer.name] = avail
                        print(day_couples)
                    if day_obj:
                        cawt[day_obj.id].append(day_couples)
                        cawt_with_group_lessons[day_obj.id].append(day_couples1)


            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
            # Merge all couples from all groups into a single window per day        
            for day_id in cawt:
                merged_couples = {}
                for couples_dict in cawt[day_id]:
                    merged_couples.update(couples_dict)
                cawt[day_id] = [merged_couples]
            for day_id in cawt_with_group_lessons:
                merged_couples_lessons = {}
                for couples_dict in cawt_with_group_lessons[day_id]:
                    merged_couples_lessons.update(couples_dict)
                cawt_with_group_lessons[day_id] = [merged_couples_lessons]
            schedule_for_these_days = []
            if forday == 'all':
                schedule_for_these_days = Day.objects.filter(user=request.user, club=club).prefetch_related(
                    'trainers__group_lesson',
                    'trainers__day_availabilities',
                    'couples__group',
                    'group_lessons__groups__couples'
                ).all()
            else:
                schedule_for_these_days.append(
                    Day.objects.filter(user=request.user, club=club).prefetch_related(
                        'trainers__group_lesson',
                        'trainers__day_availabilities',
                        'couples__group',
                        'group_lessons__groups__couples'
                    ).get(name=forday)
                )
            all_schedules = {}
            for day in schedule_for_these_days:
                if day.id not in cawt or not cawt[day.id]:
                    logger.warning(f'Keine Daten für {day.name}, wird übersprungen.')
                    continue
                if day.id not in cawt_with_group_lessons or not cawt_with_group_lessons[day.id]:
                    logger.warning(f'Keine Gruppenlektion-Daten für {day.name}.')
                    continue
                trainers = day.trainers.all()

                if not trainers.exists():
                    logger.warning(f'No trainers configured for {day.name}, skipping.')
                    continue

                
                dancers = day.dancers.all()
                # for now, maybe will be changed, strange logic
                for d in dancers:
                    if d.in_couple:
                        if d.sex == "male":
                            c = get_object_or_404(Couple, man=d)
                        elif d.sex == "female":
                            c = get_object_or_404(Couple, woman=d)
                        if c not in day.couples.all() and d in day.dancers.all():
                            new_couple, n = Couple.objects.get_or_create(name=d.name, club=d.club, group=d.group)
                    elif d in day.dancers.all():
                        new_couple, n = Couple.objects.get_or_create(name=d.name, club=d.club, group=d.group)
                    day.couples.add(new_couple)
                    day.save()

                couples = day.couples.all()
                if not couples.exists():
                    logger.warning(f'No couples configured for {day.name}, skipping.')
                    continue

                try:
                    
                    tw = make_trainers_windows(trainers, day)
                    for trainer, windows in tw.items():
                        available_slots = sum(1 for _, is_avail in windows if is_avail)
                except Exception as e:
                    logger.error(f'Error creating trainer windows for {day.name}: {e}', exc_info=True)
                    continue
                if sort_couples_by == 'group':
                    sorted_couples = sort_couples_by_group(couples, Group.objects.filter(user=request.user, club=club))
                else:
                    sorted_couples = list(couples)
                optimal_timeout = min(BASE_TIMEOUT + len(sorted_couples) * TIMEOUT_PER_COUPLE, MAX_TIMEOUT)
                # Try with escalating pairing strategy
                logger.info(cawt_with_group_lessons)
                # Added 11.5
                cawt_gl = cawt_with_group_lessons.get(day.id, cawt.get(day.id, [{}]))
                if not cawt_gl:
                    cawt_gl = cawt.get(day.id, [{}])
                # Added 11.5
                cawt_data = cawt_gl[0]
                sorted_couples = [c for c in sorted_couples if c.name in cawt_data]
                
                if not sorted_couples:
                    continue

                schedule, diagnostics = create_schedule_with_escalating_pairs(
                    cawt=cawt_data, 
                    trainers_windows=tw, 
                    couples=sorted_couples, 
                    day=day,
                    timeout_seconds=optimal_timeout,
                    max_pair_count=MAX_PAIR_COUNT_DEFAULT  
                )
                
                if not schedule:
                    logger.info(f'Trying without group lessons for {day.name}...')

                    schedule, diagnostics = create_schedule_with_escalating_pairs(
                        cawt=cawt[day.id][0], 
                        trainers_windows=tw, 
                        couples=sorted_couples, 
                        day=day,
                        timeout_seconds=optimal_timeout,
                        max_pair_count=MAX_PAIR_COUNT_DEFAULT
                    )
                    
                if schedule:
                    all_schedules[day.name] = schedule
                    pairing_info = ""
                    if diagnostics.get('pairs_used', 0) > 0:
                        pairs = diagnostics.get('paired_couples', [])
                        pair_names = ", ".join([f"{p[0]} + {p[1]}" for p in pairs])
                        pairing_info = f" [Paired {diagnostics['pairs_used']}: {pair_names}]"
                    logger.info(f'✓ Schedule created for {day.name} - {diagnostics["scheduled_count"]}/{diagnostics["total_couples"]} couples in {diagnostics["time_taken"]}s ({diagnostics["iterations"]} iterations){pairing_info}')
                else:
                    logger.error(f'✗ Could not create schedule for {day.name}. Diagnostics: {diagnostics}')
                    schedule, diagnostics = build_schedule2(
                        cawt=cawt_with_group_lessons[day.id][0],
                        trainers_windows=tw,
                        couples=sorted_couples,
                        day=day,
                        hard_timeout=optimal_timeout
                    )
                    if schedule:
                        all_schedules[day.name] = schedule
                        pairing_info = ""
                        logger.info(f'✓ Schedule created for {day.name} - {diagnostics["scheduled_count"]}/{diagnostics["total_couples"]} couples {diagnostics["used_backtracking"]} unscheduled{diagnostics["unscheduled"]}')
            if not all_schedules:
                return JsonResponse({
                    'status':'error',
                    'message':f'Could not create schedule for any configured day.'
                })

            formatted_schedule = {}
            for day_name, day_schedule in all_schedules.items():
                lessons = []
                for couple, (trainer, time_str) in day_schedule.items():
                    lessons.append({
                        'couple':couple.name,
                        'trainer':trainer.name,
                        'time':time_str,
                        'duration':couple.min_duration,
                    })
                # Sort lessons by time chronologically
                lessons.sort(key=lambda x: tuple(map(int, x['time'].split(':'))))
                formatted_schedule[day_name] = lessons
            
            return JsonResponse({
                'status':'success',
                'message':'Schedule created successfully',
                'schedule': formatted_schedule
            })
        except Exception as e:
            logger.error(f'Unexpected error in create_schedule: {e}', exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': f'Server error: {str(e)}'
            }, status=500)
    return JsonResponse({
        'status':'error',
        'message':'Invalid request method'
    }, status=400)
