import pandas as pd
from datetime import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from collections import defaultdict

#days = [ "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday", "pondelok", "utorok", "streda", "štvrtok", "piatok", "sobota", "nedeľa" ]
#df = pd.read_excel('IT MOŽNOSTI.xlsx', sheet_name='K1')
#df.dropna(how='all')
#df.dropna(axis=1, how='all')

def get_hex_color(cell):
    color = cell.fill.fgColor

    # 1. Ak je farba priamo v rgb (string)
    if isinstance(color.rgb, str):
        return color.rgb.upper()

    # 2. Ak je farba objekt typu Color a má atribút rgb
    if isinstance(color, Color) and isinstance(color.rgb, str):
        return color.rgb.upper()

    # 3. Ak je farba v .value (niekedy býva hex)
    if isinstance(color.value, str):
        return color.value.upper()

    # 4. Ak je farba theme/indexed → openpyxl to nevie preložiť
    return None



def is_shade_of_white(cell):
    raw = get_hex_color(cell)
    if raw is None:
        return False

    # odstránenie alfa kanála (napr. 00FFFFFF)
    hex_color = raw[-6:]

    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    except:
        return False

    # definícia "svetlej farby"
    return r > 200 and g > 200 and b > 200
def read_dancers_availability(group, excel_file):
    wb = load_workbook(excel_file)
    ws = wb[group]

    min_row = None
    max_row = 0
    min_col = None
    max_col = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ''):
                r = cell.row
                c = cell.column

                if min_row is None or r < min_row:
                    min_row = r
                if min_col is None or c < min_col:
                    min_col = c

                if r > max_row:
                    max_row = r
                if c > max_col:
                    max_col = c
    cell_headers = [None, "Days", "Dni", "Times", "Časy"] # Headers of the columns, that shouldn't be part od dancers availability names
    colors = []
    times = [cell.value for cell in ws[get_column_letter(min_col+1)] if isinstance(cell.value, time)]
    dancers = [cell.value for cell in ws[min_row] if cell.value not in cell_headers]
    days = [cell.value for cell in ws[get_column_letter(min_col)] if cell.value != None]

    for row in ws.iter_rows(min_row=min_row+1, max_row = max_row, min_col = min_col + 2, max_col = max_col):
        row_colors = []
        for cell in row:
            color = cell.fill.fgColor.rgb
            if is_shade_of_white(cell) or color in (None, '00000000', '00FFFFFF'):
                row_colors.append(False)
            else:
                row_colors.append(True)
        colors.append(row_colors)
        
    #divide the times into days
    day_times = defaultdict(list)
    day_avail = {}
    day_dancers_avail = {}
    i = 0
    for day in days:
        while i < len(times):
            t = times[i]
            try:
                if day_times[day][0] < t < day_times[day][-1]:
                    break
            except:
                pass
            if t not in day_times[day] or t > day_times[day][-1]:
                day_times[day].append(t)
                i += 1
            else:
                break

    for day in days:
        day_avail[day] = colors[:len(day_times[day])]
        colors = colors[len(day_times[day])+1:]
        dancers_availability = defaultdict(list)
        for x in day_avail[day]:
            for i in range(len(dancers)):
                dancers_availability[dancers[i]].append(True if x[i] else False)
        day_dancers_avail[day] = dancers_availability
    
    return day_times, day_dancers_avail

if __name__ == '__main__':
    day_times, day_dancers_avail = read_dancers_availability('K2', 'IT MOŽNOSTI.xlsx')
    for day in day_times:
        print(day, day_times[day], day_dancers_avail[day],sep='\n___________\n')
